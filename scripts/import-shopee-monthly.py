"""
import-shopee-monthly.py — 有物 ERP Stage A Phase 3 資料回補 + 未來「一句話半自動抓」匯入端

功能：
1. 解密＋解析蝦皮月度訂單 xlsx（AES 加密），匯入 sales_orders + sales_order_items
2. 解析蝦皮廣告 CSV（週報／月報格式皆支援），匯入 ad_costs
3. 商品對應：沿用 remap_all.py 驗證過的規則（PRODUCT_RULES + COLOR_MAP + SIZE_PATTERNS）
4. 冪等：訂單編號（order_number）已存在 → 整單 skip；廣告費依期間起始日（ad_date）已存在 → skip
5. 已取消單 trigger 盲點處理：訂單匯入時若原始狀態已是終態取消（僅「不成立」，見下方判斷理由），
   先用暫存狀態插入（不觸發回沖）→ items 正常插入（AFTER INSERT trigger 記 sale_ship）→
   再 UPDATE 回真實狀態（AFTER UPDATE trigger 觸發 return_reversal 回沖），淨庫存效果歸零。
   （選用此法而非「product_id 留空」，因為 product_id 留空會讓該筆訂單的商品從未被庫存系統看見，
   不利未來月度盤點對帳；本法完整走過既有已驗證的 trigger 路徑，Phase 2 五段自測已驗證此模式。）

環境變數（必須，禁止硬編碼）：
  DATABASE_URL          postgresql://user:password@host:port/dbname?sslmode=require
  SHOPEE_XLSX_PASSWORD  蝦皮訂單 xlsx 解密密碼（賣家中心「買家個資保護計畫」自設）

用法：
  python scripts/import-shopee-monthly.py --orders-dir <dir> --ads-dir <dir> [--dry-run]

--dry-run：全程執行但最後 ROLLBACK，不寫入任何資料，僅印出會發生的動作與統計（安全預覽）。
"""
import argparse
import csv
import glob
import io
import os
import re
import sys
from datetime import datetime

import msoffcrypto
import openpyxl
import psycopg2

# ============================================================
# 商品對應規則（沿用 remap_all.py 已驗證邏輯，1607/1607 筆歷史資料 100% 對應過）
# ============================================================
PRODUCT_RULES = [
    (['百摺', '百褶', '風琴摺'], [], '風琴摺衛生紙套'),
    (['棉麻'], [], '日式棉麻衛生紙套'),
    (['硅藻', '矽藻', '珪藻'], [], '硅藻泥异形杯垫'),
    (['聖誕樹', '隔熱墊'], [], '實木杯墊'),
    (['實木杯墊'], [], '實木杯墊'),
    (['日系實木杯墊'], [], '實木杯墊'),
    (['日系 實木杯墊'], [], '實木杯墊'),
    (['榉木', '櫸木'], ['收納盤', '飾品', '首飾'], '實木杯墊'),
    (['胡桃木'], ['掛勾', '掛鉤', '吸盤'], '實木杯墊'),
    (['化妝鏡', '梳妝鏡'], [], '雲朵化妝鏡'),
    (['不鏽鋼雙層', '不鏽鋼 雙層置物架'], [], '不鏽鋼多层置物架'),
    (['金屬皮革', '皮革雙層', '雙層收納架', '雙層 收納架'], ['不鏽鋼'], '金屬皮革雙層收納架'),
    (['實木 收納置物架', '木質 收納置物架', '日系實木收納', '日系 木質 收納'], [], '原木三層展示收納架｜桌面香氛／飾品／文具陳列架'),
    (['收納置物架'], ['鐵藝', '金屬'], '原木三層展示收納架｜桌面香氛／飾品／文具陳列架'),
    (['實木飾品收納', '實木質感飾品'], [], '實木飾品收納盤'),
    (['鐵藝收納籃', '鐵藝 收納籃', '日韓鐵藝'], [], '韓系金屬網格收納籃（大／小）'),
    (['金屬網格', '收納籃'], [], '韓系金屬網格收納籃（大／小）'),
    (['金屬收納籃'], [], '韓系金屬網格收納籃（大／小）'),
    (['法式擦手巾', '法式 擦手巾', '奶油風掛式擦手巾'], [], '質感可愛奶油風掛式擦手巾'),
    (['擦手巾', '擦手毛巾'], ['猫咪', '貓咪'], '質感可愛奶油風掛式擦手巾'),
    (['猫咪擦手巾', '貓咪擦手巾'], [], '猫咪擦手巾'),
    (['抹布', '菠萝', '菠蘿'], [], '超細纖維菠萝格抹布'),
    (['蛋殼陶瓷', '陶瓷小缽'], [], '日式粗陶陶瓷戒指盘首饰盒小盘子小碟子'),
    (['陶瓷首飾盤', '陶瓷 首飾盤', '霧面陶瓷', '飾品收納盤'], ['蛋殼', '小缽', '實木', '鐵藝'], '韓系不規則造型陶瓷飾品收納盤'),
    (['戒指盘', '戒指盤', '首饰盒', '首飾盒'], [], '日式粗陶陶瓷戒指盘首饰盒小盘子小碟子'),
    (['雲朵玻璃', '雲朵燈', '漢堡燈', '雲朵床頭', '包豪斯雲朵'], [], '包豪斯雲朵床頭檯燈'),
    # 2026-07-07 修正：原本 OR「雲朵」或「小夜燈」任一命中即可，導致「量子小夜燈」被錯配到本商品
    # （量子系列本身有獨立規則，見下方「量子」規則，但排在此規則之後檢查不到）。
    # 收窄為僅「雲朵」命中才算，「小夜燈」單獨出現不足以判定為本商品。
    (['雲朵'], ['化妝鏡'], '包豪斯雲朵床頭檯燈'),
    (['量子', '小夜燈'], [], '極簡光雕玻璃氛圍燈｜USB充電觸控小夜燈（雙光型）'),
    (['光雕', '玻璃燈'], ['蘑菇', '雲朵'], '極簡光雕玻璃氛圍燈｜USB充電觸控小夜燈（雙光型）'),
    (['磁吸', '夜燈', '氛圍'], ['雲朵', '開瓶器', '冰箱貼'], '極簡光雕玻璃氛圍燈｜USB充電觸控小夜燈（雙光型）'),
    (['蘑菇', '花苞燈'], [], '觸摸蘑菇檯燈'),
    (['蘑菇燈'], [], '觸摸蘑菇檯燈'),
    (['花苞燈'], [], '觸摸蘑菇檯燈'),
    (['包豪斯'], ['雲朵'], '觸摸蘑菇檯燈'),
    (['擠牙膏', '牙膏器'], [], '金屬旋轉式擠牙膏器'),
    (['牙刷架', '幸運草', '彈簧'], [], '不鏽鋼幸運草造型牙刷架'),
    (['吸盤掛鉤', '吸盤掛勾', '吸盤式挂鈎', '吸盤 掛勾'], [], '吸盤掛鉤'),
    (['長尾夾', '不鏽鋼掛勾'], [], '吸盤掛鉤'),
    (['拖鞋架', '掛鞋架', '拖鞋掛'], [], '浴室拖鞋架免打孔壁掛架'),
    (['皺褶杯', '丹麥'], [], '丹麥皺褶杯'),
    (['EVA', '拖鞋'], [], 'EVA拖鞋'),
    (['踩屎感', '拖鞋'], [], 'EVA拖鞋'),
    (['補寄', '漏寄', '錯寄', '售後'], [], None),  # 明確略過（非商品，補寄/售後項目）
]

COLOR_MAP = {
    '淡雅灰': '淡雅灰', '灰': '灰', '深灰': '深灰', '淺灰': '淺灰', '淺灰色': '淺灰色',
    '槍灰色': '槍灰色', '槍灰': '槍灰色',
    '白': '白', '奶油白': '奶油白', '白色': '白色',
    '霧面黑': '霧面黑', '霧感黑': '黑色', '黑': '黑', '黑色': '黑色',
    '黃色': '黃色', '黃': '黃色', '檸檬黃': '檸檬黃',
    '綠': '綠色', '綠色': '綠色', '酪梨綠': '酪梨綠',
    '焦糖': '焦糖色', '焦糖橙': '焦糖橙', '焦糖色': '焦糖色',
    '可可棕': '可可棕', '咖啡棕': '咖啡棕', '棕色': '棕色', '咖啡': '咖啡棕',
    '橘色': '橘色', '橘': '橘色',
    '藍色': '藍色', '藍': '藍色',
    '酒紅': '酒紅色', '酒紅色': '酒紅色',
    '銀灰': '銀灰色', '銀灰色': '銀灰色', '銀色': '銀色',
    '粉色': '粉色', '粉': '粉色',
    '紅色': '紅色', '紅': '紅色',
    '卡其': '卡其色', '卡其色': '卡其色',
    '伯爵奶茶': '伯爵奶茶', '巧克力牛奶': '巧克力牛奶',
    '抹茶拿鐵': '抹茶拿鐵', '活力柳橙': '活力柳橙', '美式咖啡': '美式咖啡',
    '奶油黃': '奶油黃',
    '胡桃木': '胡桃木色', '胡桃木色': '胡桃木色',
    '編織斜紋': '編織斜紋', '斜紋編織': '編織斜紋',
    '編織粗紋': '編織粗紋', '亞麻編織': '亞麻編織',
    '可愛櫻桃': '可愛櫻桃', '法式生活': '法式生活',
}

SIZE_PATTERNS = [
    (r'(\d{2})-(\d{2})\s*\(', lambda m: f'{m.group(1)}-{m.group(2)}'),
    (r'(\d{2})-(\d{2})', lambda m: f'{m.group(1)}-{m.group(2)}'),
    (r'大款', lambda m: '大'),
    (r'小款', lambda m: '小'),
    (r'大彈簧', lambda m: '大彈簧'),
    (r'小彈簧', lambda m: '小彈簧'),
    (r'幸運草造型', lambda m: '幸運草造型'),
    (r'三格', lambda m: '三格收納盤'),
    (r'方形', lambda m: '方形收納盤'),
    (r'長方形款', lambda m: '長方形'),
]

SPECIAL_VARIANT_MAP = {
    # 圓形凹槽杯墊 系列：DB 既有變體名用簡體+舊命名（榉木／黑胡桃），與蝦皮新資料的
    # 「淺色櫸木／深色胡桃木」字面不同源，取共同字根比對（2026-07-07 Phase 3 匯入時發現並修正）
    '淺色櫸木 圓形凹槽杯墊': '榉木',
    '深色胡桃木 圓形凹槽杯墊': '胡桃',
    '淺色櫸木收納盒（可裝6個）': '淺色櫸木收納盒（可裝6個）',
    '深色胡桃木收納盒（可裝6個）': '深色胡桃木收納盒（可裝6個）',
    '可折疊': '可折疊',
    '201不鏽鋼': '201不鏽鋼',
}


def clean_variant(variant_str):
    if not variant_str:
        return ''
    v = re.sub(r'新色報到[✨❤️]*[｜|]?', '', variant_str)
    v = re.sub(r'新色出爐[✨❤️]*[｜|]?', '', v)
    v = v.strip()
    parts = v.split(',')
    return parts[0].strip()


def extract_color(variant_str):
    if not variant_str:
        return None
    for key in SPECIAL_VARIANT_MAP:
        if key in variant_str:
            return SPECIAL_VARIANT_MAP[key]
    for color_key in sorted(COLOR_MAP.keys(), key=len, reverse=True):
        if color_key in variant_str:
            return COLOR_MAP[color_key]
    return None


def extract_size(variant_str):
    if not variant_str:
        return None
    for pattern, extractor in SIZE_PATTERNS:
        m = re.search(pattern, variant_str)
        if m:
            return extractor(m)
    return None


def match_product_rule(shopee_name):
    text = shopee_name.lower()
    for keywords, excludes, prod_name in PRODUCT_RULES:
        if any(kw.lower() in text for kw in keywords):
            if excludes and any(ex.lower() in text for ex in excludes):
                continue
            return prod_name
    return '__no_match__'


def _normalize_dim(s):
    """把尺寸描述正規化到可比對的形式：去空白、統一 × / * 為 x。"""
    return re.sub(r'\s+', '', s or '').lower().replace('×', 'x').replace('*', 'x')


def find_best_product(prod_by_name, prod_name, color, size, variant_str):
    """
    比對邏輯（2026-07-07 Phase 3 匯入時發現舊版問題並修正）：
    舊版只用「顏色字串精確相等」或「單向 substring」比對，遇到 DB 變體命名把顏色放在
    描述尾端（如「10*10cm异形 - 淺灰」對「淺灰色」）就完全比對不到——這是本輪匯入前
    半數品項對應失敗的根因。修正：顏色比對前先試「去尾字色」再比一次（淺灰色→淺灰）；
    同色候選有多筆時，用尺寸/規格數字（如 30x30 vs 40x40）進一步收斂，避免誤配到同色不同
    尺寸的錯誤商品（會導致該尺寸庫存誤扣）。
    """
    if prod_name not in prod_by_name:
        return None
    variants = prod_by_name[prod_name]  # var_key(小寫) -> product_id

    def candidates_by_color(c):
        if not c:
            return []
        cl = c.lower()
        exact = [pid for vk, pid in variants.items() if vk == cl]
        if exact:
            return exact
        partial = [pid for vk, pid in variants.items() if vk and (cl in vk or vk in cl)]
        if partial:
            return partial
        if cl.endswith('色'):
            stripped = cl[:-1]
            partial2 = [pid for vk, pid in variants.items() if vk and stripped in vk]
            if partial2:
                return partial2
        return []

    cands = candidates_by_color(color)
    if len(cands) == 1:
        return cands[0]
    if len(cands) > 1:
        dims = re.findall(r'\d+\s*[×xX*]\s*\d+', variant_str or '')
        if dims:
            norm_dims = [_normalize_dim(d) for d in dims]
            refined = [pid for vk, pid in variants.items()
                       if pid in cands and any(nd in _normalize_dim(vk) for nd in norm_dims)]
            if refined:
                return refined[0]
        return cands[0]  # 無法用尺寸收斂，退回第一個顏色候選（優於完全不匹配）

    if size:
        sl = size.lower()
        size_cands = [pid for vk, pid in variants.items() if vk and sl in vk]
        if size_cands:
            return size_cands[0]

    if variant_str:
        vl = variant_str.lower()
        vs_cands = [pid for vk, pid in variants.items() if vk and (vk in vl or vl in vk)]
        if vs_cands:
            return vs_cands[0]

    if len(variants) == 1:
        return list(variants.values())[0]
    return None


def resolve_product_id(cur, prod_by_name, shopee_product_name, shopee_variant_name, created_log):
    """
    回傳 (product_id 或 None, 原因字串)。
    規則命中但既有 variant 都比不上時，沿用 remap_all.py 已驗證過的「自動建立缺少的 variant」
    做法（歷史匯入曾用此法自動建立 37 筆新款式，見 有物ERP_資料來源.md）——因為 rule 已確認
    商品家族，只是這個顏色/款式還沒建檔，比起放著不追蹤庫存，比照既有姊妹款複製成本/售價
    基準新增一筆更貼近事實。全然對不上任何商品家族（no_rule_match）才留空，不用臆測硬猜。
    """
    prod_name = match_product_rule(shopee_product_name)
    if prod_name is None:
        return None, 'skip_rule (補寄/漏寄/售後)'
    if prod_name == '__no_match__':
        return None, 'no_rule_match'
    cleaned = clean_variant(shopee_variant_name)
    color = extract_color(cleaned)
    size = extract_size(cleaned)
    pid = find_best_product(prod_by_name, prod_name, color, size, cleaned)
    if pid:
        return pid, 'matched'

    variant_label = color or size or (cleaned[:30] if cleaned else 'default')
    variants = prod_by_name.get(prod_name, {})
    base = {}
    if variants:
        base_pid = list(variants.values())[0]
        cur.execute("""SELECT category, selling_price, unit_cost_ntd, purchase_price_cny,
                              platform_fee_rate, weight_kg
                       FROM products WHERE id = %s;""", (base_pid,))
        row = cur.fetchone()
        if row:
            base = dict(zip(
                ['category', 'selling_price', 'unit_cost_ntd', 'purchase_price_cny',
                 'platform_fee_rate', 'weight_kg'], row))

    sku_base = f"YW-{prod_name[:10]}-{variant_label[:10]}"
    sku = sku_base
    n = 1
    while True:
        cur.execute("SELECT 1 FROM products WHERE sku = %s;", (sku,))
        if not cur.fetchone():
            break
        n += 1
        sku = f"{sku_base}-{n}"

    cur.execute("""
        INSERT INTO products (product_name, variant_name, sku, category, selling_price,
                               unit_cost_ntd, purchase_price_cny, platform_fee_rate, weight_kg,
                               product_status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'常駐')
        RETURNING id;
    """, (prod_name, variant_label, sku, base.get('category'), base.get('selling_price'),
          base.get('unit_cost_ntd'), base.get('purchase_price_cny'), base.get('platform_fee_rate'),
          base.get('weight_kg')))
    new_id = cur.fetchone()[0]
    prod_by_name.setdefault(prod_name, {})[variant_label.strip().lower()] = new_id
    created_log.append((prod_name, variant_label, sku, new_id))
    return new_id, f'auto_created (variant={variant_label!r})'


# ============================================================
# 蝦皮訂單 xlsx 解密 + 解析
# ============================================================
def decrypt_xlsx_to_stream(path, password):
    with open(path, 'rb') as fp:
        office_file = msoffcrypto.OfficeFile(fp)
        office_file.load_key(password=password)
        decrypted = io.BytesIO()
        office_file.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted


def parse_order_xlsx(stream, source_file):
    wb = openpyxl.load_workbook(stream, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    required = ['訂單編號', '訂單狀態', '買家帳號', '訂單成立日期', '商品總價',
                '成交手續費', '蝦皮補助運費', '其他服務費', '金流與系統處理費',
                '賣家負擔優惠券', '賣家負擔蝦幣回饋券', '蝦皮負擔優惠券',
                '商品名稱', '商品選項名稱', '商品活動價格', '數量']
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f'{source_file} 缺少必要欄位: {missing}')

    orders = {}  # order_number -> {header fields..., items: [...]}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx['訂單編號']]:
            continue
        onum = str(row[idx['訂單編號']]).strip()
        if onum not in orders:
            order_date_raw = row[idx['訂單成立日期']]
            order_date = order_date_raw.date() if isinstance(order_date_raw, datetime) else order_date_raw
            # 賣家負擔優惠券 + 賣家負擔蝦幣回饋券 併入 seller_coupon（spec 外決定，見交付摘要）
            seller_coupon = float(row[idx['賣家負擔優惠券']] or 0) + float(row[idx['賣家負擔蝦幣回饋券']] or 0)
            orders[onum] = {
                'order_number': onum,
                'raw_status': (row[idx['訂單狀態']] or '').strip(),
                'buyer_name': row[idx['買家帳號']],
                'order_date': order_date,
                'order_amount': float(row[idx['商品總價']] or 0),
                'transaction_fee': float(row[idx['成交手續費']] or 0),
                'free_shipping_subsidy': float(row[idx['蝦皮補助運費']] or 0),
                'extended_prep_fee': float(row[idx['其他服務費']] or 0),
                'payment_processing_fee': float(row[idx['金流與系統處理費']] or 0),
                'seller_coupon': seller_coupon,
                'platform_coupon': float(row[idx['蝦皮負擔優惠券']] or 0),
                'items': [],
                'source_file': source_file,
            }
        qty = int(row[idx['數量']] or 0)
        unit_price = float(row[idx['商品活動價格']] or 0)
        orders[onum]['items'].append({
            'product_name': row[idx['商品名稱']],
            'variant_name': row[idx['商品選項名稱']],
            'qty': qty,
            'unit_price': unit_price,
            'subtotal': round(qty * unit_price, 2),
        })
    return orders


# 終態取消判斷：僅「不成立」精準比對。
# 理由：資料實測發現「已完成(鑑賞期內)，但買家仍可在鑑賞期 XXXX 前申請退貨/退款。」這類狀態字串
# 同時包含「退貨」「退款」關鍵字，若用 004 trigger 的寬鬆 regex 判斷會被誤判成取消——
# 但這類單其實是「已完成」的子狀態（只是提醒買家鑑賞期未過），從未取消。
# 故此處刻意收窄為精確字串比對，避免誤傷正常已完成訂單的庫存記帳。
TERMINAL_CANCEL_STATUSES = {'不成立'}


def is_terminal_cancel(raw_status):
    return raw_status in TERMINAL_CANCEL_STATUSES


# ============================================================
# 廣告 CSV 解析（週報／月報格式皆支援；欄位用表頭名稱查找，不寫死欄位序號）
# ============================================================
def parse_ad_csv(path):
    with open(path, encoding='utf-8-sig', newline='') as fp:
        reader = list(csv.reader(fp))
    period_row = reader[5]
    period = period_row[1] if len(period_row) > 1 else ''
    m = re.match(r'\s*(\d{4}/\d{2}/\d{2})\s*-\s*(\d{4}/\d{2}/\d{2})\s*', period)
    if not m:
        raise ValueError(f'{path} 期間格式無法解析: {period!r}')
    start_str, end_str = m.group(1), m.group(2)
    start_date = datetime.strptime(start_str, '%Y/%m/%d').date()
    header_row = reader[7]
    if '花費' not in header_row:
        raise ValueError(f'{path} 找不到「花費」欄位，header={header_row}')
    spend_idx = header_row.index('花費')
    total = 0.0
    for r in reader[8:]:
        if not r or not (r[0] or '').strip():
            continue
        try:
            total += float(r[spend_idx])
        except (ValueError, IndexError):
            continue
    description = f'蝦皮廣告 {start_str} - {end_str}'
    return start_date, round(total, 2), description


# ============================================================
# 主流程
# ============================================================
def build_product_index(cur):
    cur.execute("SELECT id, product_name, variant_name FROM products;")
    prod_by_name = {}
    for pid, name, variant in cur.fetchall():
        var_key = (variant or '').strip().lower()
        prod_by_name.setdefault(name, {})[var_key] = pid
    return prod_by_name


def import_orders(conn, orders_dir, xlsx_password, dry_run):
    files = sorted(glob.glob(os.path.join(orders_dir, 'Order.all.*.xlsx')))
    if not files:
        print(f'[警告] {orders_dir} 下找不到 Order.all.*.xlsx')
        return

    cur = conn.cursor()
    prod_by_name = build_product_index(cur)

    stats = {'files': {}, 'total_orders_seen': 0, 'total_orders_inserted': 0,
              'total_orders_skipped_existing': 0, 'total_items_inserted': 0,
              'total_items_unmapped': 0, 'cancel_orders_reversed': 0}
    unmapped_log = []
    created_log = []

    for f in files:
        fname = os.path.basename(f)
        print(f'\n=== 處理 {fname} ===')
        stream = decrypt_xlsx_to_stream(f, xlsx_password)
        orders = parse_order_xlsx(stream, fname)
        seen = len(orders)
        inserted = 0
        skipped_existing = 0
        items_inserted = 0
        items_unmapped = 0

        for onum, o in orders.items():
            cur.execute("SELECT id FROM sales_orders WHERE order_number = %s;", (onum,))
            if cur.fetchone():
                skipped_existing += 1
                continue

            terminal_cancel = is_terminal_cancel(o['raw_status'])
            net_revenue = round(
                o['order_amount'] - o['transaction_fee'] - o['extended_prep_fee']
                - o['payment_processing_fee'] - o['seller_coupon'] - o['platform_coupon'], 2)
            insert_status = 'importing_temp' if terminal_cancel else o['raw_status']

            cur.execute("""
                INSERT INTO sales_orders
                    (order_number, order_date, buyer_name, order_amount, transaction_fee,
                     free_shipping_subsidy, extended_prep_fee, payment_processing_fee,
                     seller_coupon, platform_coupon, net_revenue, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id;
            """, (onum, o['order_date'], o['buyer_name'], o['order_amount'], o['transaction_fee'],
                  o['free_shipping_subsidy'], o['extended_prep_fee'], o['payment_processing_fee'],
                  o['seller_coupon'], o['platform_coupon'], net_revenue, insert_status))
            order_id = cur.fetchone()[0]

            for item in o['items']:
                pid, reason = resolve_product_id(cur, prod_by_name, item['product_name'],
                                                  item['variant_name'], created_log)
                if pid is None:
                    items_unmapped += 1
                    unmapped_log.append((onum, item['product_name'][:50], item['variant_name'], reason))
                cur.execute("""
                    INSERT INTO sales_order_items
                        (order_id, product_id, product_name, variant_name, qty, unit_price, subtotal)
                    VALUES (%s,%s,%s,%s,%s,%s,%s);
                """, (order_id, pid, item['product_name'], item['variant_name'],
                      item['qty'], item['unit_price'], item['subtotal']))
                items_inserted += 1

            if terminal_cancel:
                cur.execute("UPDATE sales_orders SET status = %s WHERE id = %s;",
                            (o['raw_status'], order_id))
                stats['cancel_orders_reversed'] += 1

            inserted += 1

        stats['files'][fname] = {'seen': seen, 'inserted': inserted, 'skipped_existing': skipped_existing,
                                   'items_inserted': items_inserted, 'items_unmapped': items_unmapped}
        stats['total_orders_seen'] += seen
        stats['total_orders_inserted'] += inserted
        stats['total_orders_skipped_existing'] += skipped_existing
        stats['total_items_inserted'] += items_inserted
        stats['total_items_unmapped'] += items_unmapped
        print(f'  訂單: 原始 {seen} 筆 / 新增 {inserted} 筆 / 已存在跳過 {skipped_existing} 筆')
        print(f'  品項: 新增 {items_inserted} 筆 / 未對應商品 {items_unmapped} 筆')

    cur.close()
    return stats, unmapped_log, created_log


def import_ads(conn, ads_dir):
    files = sorted(glob.glob(os.path.join(ads_dir, 'Ads.overall.*.csv')))
    if not files:
        print(f'[警告] {ads_dir} 下找不到 Ads.overall.*.csv')
        return {'seen': 0, 'inserted': 0, 'skipped_existing': 0}

    cur = conn.cursor()
    stats = {'seen': 0, 'inserted': 0, 'skipped_existing': 0}
    print(f'\n=== 處理廣告 CSV ({len(files)} 份) ===')
    for f in files:
        stats['seen'] += 1
        ad_date, amount, description = parse_ad_csv(f)
        cur.execute("SELECT id FROM ad_costs WHERE ad_date = %s;", (ad_date,))
        if cur.fetchone():
            stats['skipped_existing'] += 1
            print(f'  [skip 已存在] {os.path.basename(f)} -> {description} (${amount})')
            continue
        cur.execute("""
            INSERT INTO ad_costs (ad_date, amount, description) VALUES (%s, %s, %s);
        """, (ad_date, amount, description))
        stats['inserted'] += 1
        print(f'  [新增] {os.path.basename(f)} -> {description} (${amount})')
    cur.close()
    return stats


def verify_seam(conn):
    """接縫核對：查 2026-02-25~03-05 訂單數，確認無重複無漏。"""
    cur = conn.cursor()
    cur.execute("""
        SELECT to_char(order_date, 'YYYY-MM-DD'), count(*)
        FROM sales_orders
        WHERE order_date BETWEEN '2026-02-25' AND '2026-03-05'
        GROUP BY 1 ORDER BY 1;
    """)
    rows = cur.fetchall()
    cur.execute("""
        SELECT count(*) FROM (
          SELECT order_number FROM sales_orders GROUP BY order_number HAVING count(*) > 1
        ) t;
    """)
    dup_count = cur.fetchone()[0]
    cur.close()
    print('\n=== 接縫核對 2026-02-25 ~ 03-05 ===')
    for d, c in rows:
        print(f'  {d}: {c} 筆')
    print(f'  order_number 重複筆數（應為 0）: {dup_count}')
    return dup_count


def verify_cancel_net_zero(conn):
    """驗證本輪匯入的取消單，ledger 淨庫存效果為 0。"""
    cur = conn.cursor()
    cur.execute("""
        SELECT soi.order_id, SUM(il.qty_delta)
        FROM sales_order_items soi
        JOIN sales_orders so ON so.id = soi.order_id
        JOIN inventory_ledger il ON il.ref_type = 'sales_order_item' AND il.ref_no = soi.id::text
        WHERE so.status = '不成立' AND so.order_date >= '2026-03-01'
        GROUP BY soi.order_id
        HAVING SUM(il.qty_delta) != 0;
    """)
    bad = cur.fetchall()
    cur.close()
    print(f'\n=== 取消單淨庫存效果驗證（應全數為 0，異常筆數: {len(bad)}）===')
    if bad:
        for order_id, net in bad:
            print(f'  [FAIL] order_id={order_id} net={net}')
    else:
        print('  PASS：本輪匯入的「不成立」訂單，sale_ship + return_reversal 淨效果皆為 0')
    return len(bad)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--orders-dir', required=True)
    ap.add_argument('--ads-dir', required=True)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    db_url = os.environ.get('DATABASE_URL')
    xlsx_password = os.environ.get('SHOPEE_XLSX_PASSWORD')
    if not db_url:
        print('[錯誤] 缺少環境變數 DATABASE_URL'); sys.exit(1)
    if not xlsx_password:
        print('[錯誤] 缺少環境變數 SHOPEE_XLSX_PASSWORD'); sys.exit(1)

    conn = psycopg2.connect(db_url)
    try:
        order_stats, unmapped_log, created_log = import_orders(conn, args.orders_dir, xlsx_password, args.dry_run)
        ad_stats = import_ads(conn, args.ads_dir)
        dup_count = verify_seam(conn)
        cancel_fail_count = verify_cancel_net_zero(conn)

        print('\n========== 總結 ==========')
        print(f'訂單: 原始看到 {order_stats["total_orders_seen"]} 筆 / 新增 {order_stats["total_orders_inserted"]} 筆 / '
              f'已存在跳過 {order_stats["total_orders_skipped_existing"]} 筆')
        print(f'品項: 新增 {order_stats["total_items_inserted"]} 筆 / 未對應商品 {order_stats["total_items_unmapped"]} 筆')
        print(f'取消單（回沖）處理: {order_stats["cancel_orders_reversed"]} 筆')
        print(f'廣告費: 原始 {ad_stats["seen"]} 份 / 新增 {ad_stats["inserted"]} 筆 / 已存在跳過 {ad_stats["skipped_existing"]} 筆')

        if created_log:
            print(f'\n自動建立的新商品款式（{len(created_log)} 筆，沿用姊妹款成本/售價基準）：')
            for prod_name, variant_label, sku, new_id in created_log:
                print(f'  id={new_id} | {prod_name} | {variant_label} | sku={sku}')

        if unmapped_log:
            print(f'\n未對應商品明細（{len(unmapped_log)} 筆，product_id 留空，不影響庫存以外欄位）：')
            for onum, pname, vname, reason in unmapped_log:
                print(f'  訂單{onum} | {pname} | {vname} | 原因: {reason}')

        if args.dry_run:
            conn.rollback()
            print('\n[DRY-RUN] 已 ROLLBACK，未寫入任何資料。')
        else:
            if dup_count > 0 or cancel_fail_count > 0:
                conn.rollback()
                print('\n[中止] 驗證未過（重複訂單或取消單淨效果非 0），已 ROLLBACK，請檢查。')
                sys.exit(1)
            conn.commit()
            print('\n[完成] 已 COMMIT。')
    finally:
        conn.close()


if __name__ == '__main__':
    main()
